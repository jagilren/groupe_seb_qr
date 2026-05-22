#!/usr/bin/env python3
"""
excel_migrator.py
=================
Genera la estructura de archivos del sitio QR Groupe SEB a partir de
plantilla_sitio.xlsx y la escribe en docs/ (lista para GitHub Pages).

Uso:
    python3 excel_migrator.py                       # default: plantilla_sitio.xlsx → docs/
    python3 excel_migrator.py <plantilla> <output>  # rutas custom
"""

import html
import json
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────────
# Configuración del sitio
# ─────────────────────────────────────────────────────────────────────────
REPO_ROOT = Path(__file__).parent
DEFAULT_INPUT = REPO_ROOT / "plantilla_sitio.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "docs"
TAG_RESOURCES_PATH = REPO_ROOT / "TAG_RESOURCES.xlsx"
SITE_CONFIG_PATH = REPO_ROOT / "site_config.json"

# Tema por defecto (overrideable vía site_config.json -> "theme")
# Color de la barra del navegador en móvil (override desde theme["accent"] al cargar)
_MOBILE_THEME_COLOR = "#0969da"

DEFAULT_THEME = {
    "accent": "#0969da",
    "accent_hover": "#0550ae",
    "banner_gradient": "linear-gradient(135deg, #0a3d8f 0%, #0969da 60%, #2778d9 100%)",
    "banner_text": "#ffffff",
    "banner_shadow": "0 4px 14px rgba(9,105,218,0.18)",
    "content_bg": "#f5f7fa",
    "card_bg": "#ffffff",
    "text": "#1f2328",
    "text_muted": "#57606a",
    "border": "#e1e4e8",
}

# Recursos por TAG (TAG_RESOURCES.xlsx) — pares (columna, label visible en botón)
TAG_RESOURCES_COLUMNS = [
    ("Specifications", "Especificaciones"),
    ("Handbook_Manual", "Manual"),
    ("Maintenance", "Mantenimiento"),
]

SITE_TITLE = "QR Groupe SEB"
SITE_SUBTITLE = "Documentación técnica"
SITE_URL = "https://jagilren.github.io/groupe_seb_qr/"

# Banner: título centrado entre los dos logos
BANNER_TITLE_FULL = "Documentación Técnica — PTAR STARnD"
BANNER_TITLE_SHORT = "PTAR STARnD"

# Placeholders SVG inline (reemplazar por los logos reales después)
LOGO_RPCI_SVG = """<svg class="banner-logo-svg" viewBox="0 0 140 56" xmlns="http://www.w3.org/2000/svg" aria-label="RPCI">
  <rect x="2" y="2" width="136" height="52" rx="8" fill="#ffffff" stroke="#0969da" stroke-width="2"/>
  <text x="70" y="37" font-family="-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif" font-size="22" font-weight="800" letter-spacing="1.5" fill="#0969da" text-anchor="middle">RPCI</text>
</svg>"""

LOGO_CLIENT_SVG = """<svg class="banner-logo-svg" viewBox="0 0 140 56" xmlns="http://www.w3.org/2000/svg" aria-label="STARnD">
  <rect x="2" y="2" width="136" height="52" rx="8" fill="#ffffff" stroke="#198754" stroke-width="2"/>
  <text x="70" y="36" font-family="-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif" font-size="20" font-weight="800" letter-spacing="0.5" fill="#198754" text-anchor="middle">STARnD</text>
</svg>"""

# Fallback genérico para el thumbnail del botón "Ver imagen"
GENERIC_IMAGE_SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="none" stroke="#94a3b8" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round">
  <rect x="3" y="3" width="18" height="18" rx="2"/>
  <circle cx="9" cy="9" r="1.5" fill="#94a3b8"/>
  <polyline points="3 17 9 12 14 16 17 13 21 17"/>
</svg>"""

# Orden fijo de las categorías en el sitio
CATEGORY_ORDER = ["EQUIPOS", "INSTRUMENTOS", "TANQUES"]

# Columnas fijas en la plantilla (las demás son propiedades dinámicas)
FIXED_COLS_HEAD = ["TAG", "Categoría", "Título"]
FIXED_COLS_TAIL = ["Ficha Técnica", "Curva"]


# ─────────────────────────────────────────────────────────────────────────
# Data
# ─────────────────────────────────────────────────────────────────────────
@dataclass
class Item:
    tag: str
    category: str
    title: str  # default = tag si está vacío
    properties: dict[str, str] = field(default_factory=dict)
    ficha_tecnica_url: Optional[str] = None
    curva_url: Optional[str] = None
    images: list[str] = field(default_factory=list)  # nombres de archivo (1.png, 2.jpeg…)
    resources: dict[str, str] = field(default_factory=dict)  # {col_name: url} desde TAG_RESOURCES.xlsx

    @property
    def slug(self) -> str:
        return self.tag.replace(" ", "").replace("/", "_")

    @property
    def filename(self) -> str:
        return self.slug + ".html"


# ─────────────────────────────────────────────────────────────────────────
# Lectura de la plantilla
# ─────────────────────────────────────────────────────────────────────────
def cell_value(cell) -> str:
    """Convierte un valor de celda a string limpio."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def cell_hyperlink(cell) -> Optional[str]:
    """Devuelve el target del hyperlink si la celda tiene uno."""
    if cell.hyperlink is not None:
        target = cell.hyperlink.target or cell.hyperlink.location
        if target:
            return str(target).strip()
    # Fallback: si el valor de la celda parece URL
    v = cell_value(cell)
    if v.startswith(("http://", "https://")):
        return v
    return None


def load_site_config(path: Path) -> dict:
    """Lee site_config.json (opcional). Devuelve dict con defaults rellenados.

    Schema:
    {
      "site_title": "...",
      "banner_title_full": "...",
      "banner_title_short": "...",
      "site_url": "...",
      "theme": { ...DEFAULT_THEME keys... }
    }
    """
    cfg = {
        "site_title": SITE_TITLE,
        "banner_title_full": BANNER_TITLE_FULL,
        "banner_title_short": BANNER_TITLE_SHORT,
        "site_url": SITE_URL,
        "theme": dict(DEFAULT_THEME),
    }
    if not path.exists():
        print(f"   ℹ️  Sin site_config.json (usando tema por defecto)")
        return cfg
    user_cfg = json.loads(path.read_text(encoding="utf-8"))
    for k in ("site_title", "banner_title_full", "banner_title_short", "site_url"):
        if k in user_cfg:
            cfg[k] = user_cfg[k]
    if "theme" in user_cfg and isinstance(user_cfg["theme"], dict):
        cfg["theme"].update(user_cfg["theme"])
    print(f"🎨 Tema cargado de site_config.json")
    return cfg


def build_theme_override_css(theme: dict) -> str:
    """Genera un bloque :root con las variables CSS sobrescritas desde el config.

    Se appendea después de STYLES_CSS para que la cascada lo aplique encima de los defaults.
    """
    return f"""
/* ---------------- Overrides desde site_config.json ---------------- */
:root {{
    --bg: {theme['card_bg']};
    --border: {theme['border']};
    --text: {theme['text']};
    --text-muted: {theme['text_muted']};
    --accent: {theme['accent']};
    --accent-hover: {theme['accent_hover']};
    --banner-bg: {theme['banner_gradient']};
    --banner-text: {theme['banner_text']};
    --banner-shadow: {theme['banner_shadow']};
    --content-bg: {theme['content_bg']};
}}
"""


def load_tag_resources(xlsx_path: Path) -> dict[str, dict[str, str]]:
    """Lee TAG_RESOURCES.xlsx → {TAG_ID: {col_name: url}}.

    Si el archivo no existe, devuelve {} (los botones simplemente no se renderizan).
    """
    if not xlsx_path.exists():
        print(f"   ℹ️  Sin TAG_RESOURCES.xlsx (no se renderizan botones de recursos)")
        return {}

    print(f"📖 Leyendo recursos: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["TAG_RESOURCES"] if "TAG_RESOURCES" in wb.sheetnames else wb.active

    headers = [cell_value(c) for c in ws[1]]
    if not headers or headers[0] != "TAG_ID":
        raise ValueError(f"TAG_RESOURCES.xlsx: primera columna debe ser 'TAG_ID', es {headers[0]!r}")

    # Solo tomamos las columnas configuradas en TAG_RESOURCES_COLUMNS que existan
    col_idx = {}
    for col_name, _label in TAG_RESOURCES_COLUMNS:
        if col_name in headers:
            col_idx[col_name] = headers.index(col_name)

    resources: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2):
        tag = cell_value(row[0])
        if not tag:
            continue
        per_tag: dict[str, str] = {}
        for col_name, idx in col_idx.items():
            url = cell_hyperlink(row[idx])
            if url:
                per_tag[col_name] = url
        if per_tag:
            resources[tag] = per_tag

    print(f"   🔗 {len(resources)} TAGs con recursos cargados ({sum(len(v) for v in resources.values())} enlaces)")
    return resources


def load_items(xlsx_path: Path) -> tuple[list[Item], list[str]]:
    """Lee plantilla_sitio.xlsx → (items, lista de propiedades en orden)."""
    print(f"📖 Leyendo: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Datos"] if "Datos" in wb.sheetnames else wb.active

    # Headers (fila 1)
    headers = [cell_value(c) for c in ws[1]]
    if not headers or headers[0] != "TAG":
        raise ValueError(f"La primera columna debe ser 'TAG', pero es {headers[0]!r}")

    # Identificar columnas: TAG, Categoría, Título → propiedades → Ficha Técnica, Curva
    try:
        idx_tag = headers.index("TAG")
        idx_cat = headers.index("Categoría")
        idx_tit = headers.index("Título")
        idx_ft = headers.index("Ficha Técnica") if "Ficha Técnica" in headers else None
        idx_curva = headers.index("Curva") if "Curva" in headers else None
    except ValueError as e:
        raise ValueError(f"Falta columna obligatoria: {e}")

    # Columnas de propiedades = entre Título y Ficha Técnica (o hasta el final)
    prop_start = idx_tit + 1
    prop_end = idx_ft if idx_ft is not None else len(headers)
    prop_columns = [h for h in headers[prop_start:prop_end] if h]

    items: list[Item] = []
    for row in ws.iter_rows(min_row=2):
        tag = cell_value(row[idx_tag])
        if not tag:
            continue  # fila vacía
        category = cell_value(row[idx_cat]).upper()
        if category not in CATEGORY_ORDER:
            print(f"  ⚠️  TAG {tag}: categoría '{category}' no reconocida; se asigna EQUIPOS")
            category = "EQUIPOS"
        title = cell_value(row[idx_tit]) or tag

        # Propiedades: solo las que tengan valor
        props = {}
        for i, col_name in enumerate(headers[prop_start:prop_end]):
            if not col_name:
                continue
            val = cell_value(row[prop_start + i])
            if val:
                props[col_name] = val

        ficha_url = cell_hyperlink(row[idx_ft]) if idx_ft is not None else None
        curva_url = cell_hyperlink(row[idx_curva]) if idx_curva is not None else None

        items.append(Item(
            tag=tag,
            category=category,
            title=title,
            properties=props,
            ficha_tecnica_url=ficha_url,
            curva_url=curva_url,
        ))

    # Cargar manifest de imágenes (si existe) y asignarlo a cada item
    manifest_path = DEFAULT_OUTPUT / "assets" / "images" / "manifest.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        for it in items:
            if it.slug in manifest:
                it.images = manifest[it.slug]
        with_imgs = sum(1 for it in items if it.images)
        print(f"   📷 {with_imgs} TAGs con imágenes cargadas del manifest")
    else:
        print(f"   ℹ️  Sin manifest de imágenes (corre convert_fichas_to_template.py primero si las quieres)")

    # Cargar recursos por TAG (Specifications / Handbook_Manual / Maintenance)
    resources_by_tag = load_tag_resources(TAG_RESOURCES_PATH)
    for it in items:
        if it.tag in resources_by_tag:
            it.resources = resources_by_tag[it.tag]

    print(f"   {len(items)} TAGs cargados, {len(prop_columns)} propiedades posibles")
    return items, prop_columns


# ─────────────────────────────────────────────────────────────────────────
# Helpers de HTML
# ─────────────────────────────────────────────────────────────────────────
def h(s: str) -> str:
    """HTML escape."""
    return html.escape(s, quote=True)


def group_by_category(items: list[Item]) -> dict[str, list[Item]]:
    """Devuelve dict {categoría: [items ordenados por TAG]}."""
    grouped: dict[str, list[Item]] = {c: [] for c in CATEGORY_ORDER}
    for item in items:
        grouped.setdefault(item.category, []).append(item)
    for cat in grouped:
        grouped[cat].sort(key=lambda i: i.tag)
    return grouped


def build_nav(grouped: dict[str, list[Item]], rel_prefix: str, active_tag: Optional[str] = None) -> str:
    """Construye el <nav> del sidebar con paths relativos según ubicación."""
    out = ['<nav class="nav"><ul class="nav-list">']
    for cat in CATEGORY_ORDER:
        items = grouped.get(cat, [])
        if not items:
            continue
        cat_slug = cat.lower()
        out.append('<li class="nav-group">')
        out.append('<div class="nav-cat-row">')
        out.append(f'<button class="nav-cat" aria-expanded="false">{h(cat)}</button>')
        out.append('</div>')
        out.append(f'<a class="nav-cat-link" href="{rel_prefix}{cat_slug}/index.html">Ver categoría completa →</a>')
        out.append('<ul class="nav-items">')
        for item in items:
            active_cls = ' class="active"' if active_tag == item.tag else ""
            out.append(
                f'<li><a{active_cls} href="{rel_prefix}{cat_slug}/{item.filename}">{h(item.tag)}</a></li>'
            )
        out.append('</ul></li>')
    out.append('</ul></nav>')
    return "\n".join(out)


def page_skeleton(
    page_title: str,
    topbar_title: str,
    sidebar_html: str,
    content_html: str,
    rel_prefix: str,
) -> str:
    """Skeleton HTML completo (head + banner + sidebar + content)."""
    return f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, viewport-fit=cover">
    <meta name="theme-color" content="{_MOBILE_THEME_COLOR}">
    <title>{h(page_title)} — {h(SITE_TITLE)}</title>
    <link rel="stylesheet" href="{rel_prefix}styles.css">
</head>
<body data-rel="{rel_prefix}">
    <div class="sidebar-backdrop" aria-hidden="true"></div>
    <header class="banner">
        <button class="menu-toggle" aria-label="Abrir menú" aria-controls="sidebar" aria-expanded="false">☰</button>
        <a class="banner-logo banner-logo-left" href="{rel_prefix}index.html" aria-label="Inicio">
            {LOGO_RPCI_SVG}
        </a>
        <h1 class="banner-title">
            <span class="banner-title-full">{h(BANNER_TITLE_FULL)}</span>
            <span class="banner-title-short">{h(BANNER_TITLE_SHORT)}</span>
        </h1>
        <div class="banner-logo banner-logo-right" aria-label="Cliente">
            {LOGO_CLIENT_SVG}
        </div>
        <button class="search-toggle" aria-label="Buscar" aria-controls="search-modal" aria-expanded="false">
            <svg viewBox="0 0 24 24" width="22" height="22" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
                <circle cx="11" cy="11" r="7"/>
                <line x1="21" y1="21" x2="16.5" y2="16.5"/>
            </svg>
        </button>
    </header>
    <div class="lightbox" id="lightbox" hidden role="dialog" aria-label="Galería de imágenes" aria-modal="true">
        <div class="lightbox-overlay" aria-hidden="true"></div>
        <button class="lightbox-close" aria-label="Cerrar galería">✕</button>
        <button class="lightbox-prev" aria-label="Imagen anterior">‹</button>
        <button class="lightbox-next" aria-label="Imagen siguiente">›</button>
        <figure class="lightbox-figure">
            <img class="lightbox-img" alt="">
            <figcaption class="lightbox-caption"></figcaption>
        </figure>
        <div class="lightbox-counter" aria-live="polite">1 / 1</div>
    </div>
    <div class="search-modal" id="search-modal" hidden role="dialog" aria-label="Búsqueda" aria-modal="true">
        <div class="search-modal-overlay" aria-hidden="true"></div>
        <div class="search-modal-box" role="document">
            <div class="search-modal-head">
                <svg class="search-modal-icon" viewBox="0 0 24 24" width="20" height="20" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
                    <circle cx="11" cy="11" r="7"/>
                    <line x1="21" y1="21" x2="16.5" y2="16.5"/>
                </svg>
                <input type="search" class="search-input" placeholder="Buscar TAG, función, marca…" autocomplete="off" autocapitalize="off" autocorrect="off" spellcheck="false" aria-label="Buscar">
                <button class="search-modal-close" aria-label="Cerrar búsqueda">✕</button>
            </div>
            <div class="search-results" aria-live="polite"></div>
            <div class="search-hint">
                <kbd>Esc</kbd> para cerrar
            </div>
        </div>
    </div>
    <div class="layout">
        <aside class="sidebar" id="sidebar">
            <div class="brand">
                <a href="{rel_prefix}index.html">
                    <h1>{h(SITE_TITLE)}</h1>
                </a>
                <button class="sidebar-close" aria-label="Cerrar menú">✕</button>
            </div>
            {sidebar_html}
        </aside>
        <main class="content">
            <article>
{content_html}
            </article>
        </main>
    </div>
    <script src="{rel_prefix}script.js"></script>
</body>
</html>
"""


# ─────────────────────────────────────────────────────────────────────────
# Renderizado de cada tipo de página
# ─────────────────────────────────────────────────────────────────────────
def render_home(grouped: dict[str, list[Item]]) -> str:
    parts = [
        f'<header class="page-header"><h1>{h(SITE_TITLE)}</h1>',
        f'<p class="subtitle">{h(SITE_SUBTITLE)}</p></header>',
        '<div class="accordion">',
    ]
    for cat in CATEGORY_ORDER:
        items = grouped.get(cat, [])
        if not items:
            continue
        cat_slug = cat.lower()
        parts.append('<div class="acc-item">')
        parts.append(
            f'<button class="acc-header" aria-expanded="false">'
            f'<span class="acc-title">{h(cat)}</span>'
            f'<span class="acc-meta">{len(items)} elementos</span>'
            f'<span class="acc-chevron" aria-hidden="true">›</span></button>'
        )
        parts.append('<div class="acc-body">')
        parts.append('<ul class="acc-list">')
        for item in items:
            parts.append(
                f'<li><a href="{cat_slug}/{item.filename}">{h(item.tag)}</a></li>'
            )
        parts.append('</ul>')
        parts.append(f'<a class="acc-cta" href="{cat_slug}/index.html">Ver categoría completa →</a>')
        parts.append('</div></div>')
    parts.append('</div>')
    return "\n".join(parts)


def render_category_index(category: str, items: list[Item]) -> str:
    parts = [
        f'<header class="page-header"><h1>{h(category)}</h1>',
        f'<p class="breadcrumb"><a href="../index.html">Inicio</a> / <span>{h(category)}</span></p>',
        f'<p class="count">{len(items)} elementos</p></header>',
        '<div class="accordion">',
    ]
    for item in items:
        # Preview: primeras 3 propiedades con valor
        preview_rows = []
        for k, v in list(item.properties.items())[:3]:
            preview_rows.append(
                f'<div class="acc-row"><span class="acc-row-label">{h(k)}</span>'
                f'<span class="acc-row-value">{h(v)}</span></div>'
            )
        preview_html = "".join(preview_rows) if preview_rows else '<p class="empty">Sin datos previos.</p>'
        # Meta: título si difiere del TAG
        meta = h(item.title) if item.title != item.tag else ""
        parts.append('<div class="acc-item">')
        parts.append(
            f'<button class="acc-header" aria-expanded="false">'
            f'<span class="acc-title">{h(item.tag)}</span>'
            f'<span class="acc-meta">{meta}</span>'
            f'<span class="acc-chevron" aria-hidden="true">›</span></button>'
        )
        parts.append(
            f'<div class="acc-body">{preview_html}'
            f'<a class="acc-cta" href="{item.filename}">Ver ficha completa →</a></div>'
        )
        parts.append('</div>')
    parts.append('</div>')
    return "\n".join(parts)


def render_item_page(item: Item) -> str:
    parts = [
        f'<header class="page-header"><h1>{h(item.title)}</h1>',
        f'<p class="breadcrumb">'
        f'<a href="../index.html">Inicio</a> / '
        f'<a href="index.html">{h(item.category)}</a> / '
        f'<span>{h(item.tag)}</span></p>',
        '</header>',
    ]

    # Galería de imágenes (si el TAG tiene)
    if item.images:
        img_urls = [
            f"../assets/images/{item.slug}/{fname}" for fname in item.images
        ]
        # data-images: JSON con las URLs (escapar comillas)
        data_attr = json.dumps(img_urls).replace('"', '&quot;')
        n = len(item.images)
        label = "Ver imagen" if n == 1 else f"Ver {n} imágenes"
        thumb_src = img_urls[0]
        fallback = "../assets/icons/generic-image.svg"
        parts.append(
            f'<button class="gallery-btn" data-images="{data_attr}" '
            f'data-tag="{h(item.tag)}">'
            f'<img class="gallery-btn-thumb" src="{thumb_src}" alt="" '
            f'loading="lazy" '
            f'onerror="this.onerror=null;this.src=\'{fallback}\'">'
            f'<span class="gallery-btn-label">{label}</span></button>'
        )

    if item.properties:
        parts.append('<div class="accordion">')
        for label, value in item.properties.items():
            parts.append('<div class="acc-item">')
            parts.append(
                f'<button class="acc-header" aria-expanded="false">'
                f'<span class="acc-title">{h(label)}</span>'
                f'<span class="acc-chevron" aria-hidden="true">›</span></button>'
            )
            # Multiline support: dividir por saltos de línea en párrafos
            paragraphs = []
            for line in value.split("\n"):
                line = line.strip()
                if line:
                    paragraphs.append(f'<p>{h(line)}</p>')
            body = "\n".join(paragraphs) if paragraphs else '<p class="empty">—</p>'
            parts.append(f'<div class="acc-body">{body}</div></div>')
        parts.append('</div>')
    else:
        parts.append('<p class="empty">Esta ficha no tiene propiedades cargadas.</p>')

    # Documentos y enlaces (Ficha Técnica/Curva de plantilla + recursos de TAG_RESOURCES)
    resource_links = [
        (label, item.resources[col])
        for col, label in TAG_RESOURCES_COLUMNS
        if item.resources.get(col)
    ]
    if item.ficha_tecnica_url or item.curva_url or resource_links:
        parts.append('<section class="external-links">')
        parts.append('<h2>Documentos y enlaces</h2>')
        parts.append('<ul class="link-list">')
        if item.ficha_tecnica_url:
            parts.append(
                f'<li><a class="link-btn" href="{h(item.ficha_tecnica_url)}" '
                f'target="_blank" rel="noopener">Ficha Técnica ↗</a></li>'
            )
        if item.curva_url:
            parts.append(
                f'<li><a class="link-btn" href="{h(item.curva_url)}" '
                f'target="_blank" rel="noopener">Curva ↗</a></li>'
            )
        for label, url in resource_links:
            parts.append(
                f'<li><a class="link-btn" href="{h(url)}" '
                f'target="_blank" rel="noopener">{h(label)} ↗</a></li>'
            )
        parts.append('</ul></section>')

    return "\n".join(parts)


# ─────────────────────────────────────────────────────────────────────────
# Assets estáticos (CSS + JS) embebidos
# ─────────────────────────────────────────────────────────────────────────
STYLES_CSS = r""":root {
    --bg: #ffffff;
    --bg-alt: #f6f8fa;
    --border: #e1e4e8;
    --text: #1f2328;
    --text-muted: #57606a;
    --accent: #0969da;
    --accent-hover: #0550ae;
    --banner-bg: linear-gradient(135deg, #0a3d8f 0%, #0969da 60%, #2778d9 100%);
    --banner-text: #ffffff;
    --banner-shadow: 0 4px 14px rgba(9,105,218,0.18);
    --content-bg: #f5f7fa;
    --shadow: 0 1px 3px rgba(0,0,0,0.06), 0 1px 2px rgba(0,0,0,0.04);
    --shadow-hover: 0 4px 12px rgba(0,0,0,0.08);
    --radius: 8px;
    --banner-h: 72px;
    --banner-h-mobile: 60px;
}

* { box-sizing: border-box; margin: 0; padding: 0; }
html { -webkit-text-size-adjust: 100%; }

body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Oxygen, Ubuntu, sans-serif;
    background: var(--content-bg);
    color: var(--text);
    line-height: 1.55;
    -webkit-font-smoothing: antialiased;
    -webkit-tap-highlight-color: rgba(9,105,218,0.15);
}

.layout { display: grid; grid-template-columns: 280px 1fr; min-height: calc(100vh - var(--banner-h)); }

/* ---------------- Banner (corporativo, todas las páginas) ---------------- */
.banner {
    display: flex;
    align-items: center;
    gap: 16px;
    position: sticky;
    top: 0;
    z-index: 8;
    background: var(--banner-bg);
    color: var(--banner-text);
    border-bottom: 0;
    height: var(--banner-h);
    padding: 0 24px;
    box-shadow: var(--banner-shadow);
}
.banner-logo {
    display: flex;
    align-items: center;
    flex-shrink: 0;
    text-decoration: none;
}
.banner-logo-svg {
    height: 48px;
    width: auto;
    display: block;
}
.banner-title {
    flex: 1;
    text-align: center;
    font-size: 20px;
    font-weight: 700;
    color: var(--banner-text);
    letter-spacing: -0.01em;
    margin: 0;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
    text-shadow: 0 1px 2px rgba(0,0,0,0.15);
}
.banner-title-short { display: none; }

.menu-toggle {
    display: none;
    background: transparent;
    border: 0;
    width: 44px;
    height: 44px;
    font-size: 24px;
    cursor: pointer;
    border-radius: var(--radius);
    color: var(--banner-text);
    line-height: 1;
    flex-shrink: 0;
}
.menu-toggle:hover { background: rgba(255,255,255,0.15); }

/* ---------------- Search button + modal ---------------- */
.search-toggle {
    background: transparent;
    border: 0;
    width: 44px;
    height: 44px;
    cursor: pointer;
    border-radius: var(--radius);
    color: var(--banner-text);
    display: inline-flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
}
.search-toggle:hover { background: rgba(255,255,255,0.15); }
.search-toggle:focus-visible {
    outline: 2px solid var(--banner-text);
    outline-offset: 2px;
}

.search-modal[hidden] { display: none; }
.search-modal {
    position: fixed;
    inset: 0;
    z-index: 50;
    display: flex;
    align-items: flex-start;
    justify-content: center;
    padding: 80px 16px 16px;
}
.search-modal-overlay {
    position: absolute;
    inset: 0;
    background: rgba(15,23,42,0.55);
    backdrop-filter: blur(2px);
    animation: search-fade-in 0.15s ease;
}
.search-modal-box {
    position: relative;
    width: 100%;
    max-width: 640px;
    max-height: calc(100vh - 100px);
    display: flex;
    flex-direction: column;
    background: var(--bg);
    border-radius: 12px;
    box-shadow: 0 20px 60px rgba(0,0,0,0.25);
    overflow: hidden;
    animation: search-pop-in 0.18s ease;
}
@keyframes search-fade-in { from { opacity: 0; } to { opacity: 1; } }
@keyframes search-pop-in {
    from { opacity: 0; transform: translateY(-12px); }
    to { opacity: 1; transform: translateY(0); }
}

.search-modal-head {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 14px 16px;
    border-bottom: 1px solid var(--border);
}
.search-modal-icon { color: var(--text-muted); flex-shrink: 0; }
.search-input {
    flex: 1;
    border: 0;
    outline: 0;
    background: transparent;
    font: inherit;
    font-size: 16px;
    color: var(--text);
    padding: 4px 0;
    min-width: 0;
}
.search-input::placeholder { color: var(--text-muted); }
.search-modal-close {
    background: var(--bg-alt);
    border: 0;
    width: 32px;
    height: 32px;
    border-radius: 6px;
    cursor: pointer;
    color: var(--text-muted);
    font-size: 16px;
    line-height: 1;
    flex-shrink: 0;
}
.search-modal-close:hover { background: var(--border); color: var(--text); }

.search-results {
    flex: 1;
    overflow-y: auto;
    padding: 6px 8px 8px;
}
.search-result {
    display: block;
    padding: 10px 14px;
    border-radius: 8px;
    color: var(--text);
    text-decoration: none;
    transition: background 0.1s ease;
}
.search-result + .search-result { margin-top: 2px; }
.search-result:hover,
.search-result.active {
    background: rgba(9,105,218,0.10);
    color: var(--accent);
}
.search-result-head {
    display: flex;
    align-items: baseline;
    gap: 10px;
}
.search-result-tag {
    font-weight: 600;
    font-size: 15px;
}
.search-result-cat {
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    color: var(--text-muted);
}
.search-result.active .search-result-cat { color: var(--accent); }
.search-result-preview {
    font-size: 13px;
    color: var(--text-muted);
    margin-top: 2px;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}
.search-result mark {
    background: rgba(255,212,0,0.45);
    color: inherit;
    padding: 0 1px;
    border-radius: 2px;
}

.search-empty {
    padding: 24px 16px;
    text-align: center;
    color: var(--text-muted);
    font-size: 14px;
}

.search-hint {
    padding: 10px 16px;
    border-top: 1px solid var(--border);
    font-size: 12px;
    color: var(--text-muted);
    text-align: right;
}
.search-hint kbd {
    background: var(--bg-alt);
    border: 1px solid var(--border);
    border-radius: 4px;
    padding: 2px 6px;
    font-family: -apple-system, monospace;
    font-size: 11px;
    color: var(--text);
}

body.search-open { overflow: hidden; }

/* ---------------- Galería / lightbox ---------------- */
.gallery-btn {
    display: inline-flex;
    align-items: center;
    gap: 10px;
    background: var(--bg-alt);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    color: var(--accent);
    font: inherit;
    font-size: 14px;
    font-weight: 600;
    padding: 6px 14px 6px 6px;  /* asimétrico: izq compensa el thumb */
    margin-bottom: 16px;
    cursor: pointer;
    transition: background 0.12s ease, border-color 0.12s ease;
    line-height: 1.4;
    min-height: 44px;  /* alto fijo para que el thumb no haga crecer el botón */
}
.gallery-btn-thumb {
    width: 32px;
    height: 32px;
    border-radius: 4px;
    object-fit: cover;
    flex-shrink: 0;
    background: var(--bg);
    border: 1px solid var(--border);
    display: block;
}
.gallery-btn-label {
    flex-shrink: 0;
}
.gallery-btn:hover {
    background: rgba(9,105,218,0.08);
    border-color: var(--accent);
}
.gallery-btn:focus-visible {
    outline: 2px solid var(--accent);
    outline-offset: 2px;
}

.lightbox[hidden] { display: none; }
.lightbox {
    position: fixed;
    inset: 0;
    z-index: 60;
    display: flex;
    align-items: center;
    justify-content: center;
}
.lightbox-overlay {
    position: absolute;
    inset: 0;
    background: rgba(15,23,42,0.92);
    animation: search-fade-in 0.15s ease;
}
.lightbox-figure {
    position: relative;
    margin: 0;
    max-width: 92vw;
    max-height: calc(100vh - 80px);
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 10px;
    animation: search-pop-in 0.2s ease;
}
.lightbox-img {
    max-width: 92vw;
    max-height: calc(100vh - 140px);
    object-fit: contain;
    border-radius: 6px;
    box-shadow: 0 12px 40px rgba(0,0,0,0.6);
    background: #ffffff;
}
.lightbox-caption {
    color: rgba(255,255,255,0.9);
    font-size: 14px;
    font-weight: 500;
    text-align: center;
    max-width: 92vw;
}
.lightbox-close,
.lightbox-prev,
.lightbox-next {
    position: absolute;
    background: rgba(255,255,255,0.1);
    border: 0;
    color: white;
    cursor: pointer;
    transition: background 0.12s ease, transform 0.12s ease;
    z-index: 1;
    line-height: 1;
}
.lightbox-close:hover,
.lightbox-prev:hover,
.lightbox-next:hover {
    background: rgba(255,255,255,0.25);
}
.lightbox-close {
    top: 16px;
    right: 16px;
    width: 44px;
    height: 44px;
    border-radius: 50%;
    font-size: 22px;
}
.lightbox-prev,
.lightbox-next {
    top: 50%;
    transform: translateY(-50%);
    width: 52px;
    height: 52px;
    border-radius: 50%;
    font-size: 38px;
    font-weight: 300;
    display: flex;
    align-items: center;
    justify-content: center;
    padding-bottom: 4px;  /* corrección visual del chevron */
}
.lightbox-prev { left: 24px; }
.lightbox-next { right: 24px; }
.lightbox-prev:hover,
.lightbox-next:hover {
    transform: translateY(-50%) scale(1.05);
}
.lightbox.single .lightbox-prev,
.lightbox.single .lightbox-next,
.lightbox.single .lightbox-counter { display: none; }
.lightbox-counter {
    position: absolute;
    bottom: 20px;
    left: 50%;
    transform: translateX(-50%);
    background: rgba(255,255,255,0.12);
    color: white;
    padding: 6px 14px;
    border-radius: 999px;
    font-size: 13px;
    font-weight: 500;
}
body.lightbox-open { overflow: hidden; }

.sidebar-backdrop {
    display: none;
    position: fixed;
    inset: 0;
    background: rgba(0,0,0,0.45);
    z-index: 9;
    opacity: 0;
    transition: opacity 0.2s ease;
}
.sidebar-backdrop.show { display: block; opacity: 1; }

/* ---------------- Sidebar ---------------- */
.sidebar {
    background: var(--bg-alt);
    border-right: 1px solid var(--border);
    padding: 24px 0;
    overflow-y: auto;
    position: sticky;
    top: var(--banner-h);
    height: calc(100vh - var(--banner-h));
    display: flex;
    flex-direction: column;
}

.brand { padding: 0 24px 20px; border-bottom: 1px solid var(--border); }
.brand a { color: var(--text); text-decoration: none; }
.brand h1 { font-size: 18px; font-weight: 600; letter-spacing: -0.01em; }

.sidebar-close {
    display: none;
    background: transparent;
    border: 0;
    width: 40px;
    height: 40px;
    font-size: 20px;
    cursor: pointer;
    border-radius: 6px;
    color: var(--text-muted);
    line-height: 1;
}
.sidebar-close:hover { background: rgba(0,0,0,0.06); color: var(--text); }

.nav { flex: 1; padding: 12px 12px 8px; }
.nav-list { list-style: none; }
.nav-group { margin-bottom: 6px; }

.nav-cat-row {
    display: flex;
    align-items: center;
    gap: 4px;
}
.nav-cat {
    flex: 1;
    text-align: left;
    background: none;
    border: 0;
    padding: 10px 12px;
    font-size: 13px;
    font-weight: 600;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.04em;
    cursor: pointer;
    border-radius: 6px;
    font-family: inherit;
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.nav-cat::after {
    content: "›";
    transform: rotate(90deg);
    transition: transform 0.2s ease;
    font-size: 18px;
    margin-left: 8px;
    color: var(--text-muted);
}
.nav-group.open .nav-cat::after { transform: rotate(-90deg); }
.nav-cat:hover { background: rgba(9,105,218,0.08); color: var(--accent); }

.nav-cat-link {
    padding: 4px 12px 8px;
    font-size: 11px;
    color: var(--accent);
    text-decoration: none;
    display: block;
}
.nav-cat-link:hover { text-decoration: underline; }

.nav-items {
    list-style: none;
    padding-left: 6px;
    max-height: 0;
    overflow: hidden;
    transition: max-height 0.25s ease;
}
.nav-group.open .nav-items { max-height: 6000px; }
.nav-items li a {
    display: block;
    padding: 10px 12px;
    font-size: 14px;
    color: var(--text);
    text-decoration: none;
    border-radius: 6px;
    border-left: 2px solid transparent;
    margin-bottom: 2px;
}
.nav-items li a:hover { background: rgba(9,105,218,0.08); color: var(--accent); }
.nav-items li a.active {
    background: rgba(9,105,218,0.12);
    color: var(--accent);
    border-left-color: var(--accent);
    font-weight: 600;
}

.sidebar-footer {
    padding: 16px 24px;
    border-top: 1px solid var(--border);
    font-size: 12px;
}
.sidebar-footer a { color: var(--text-muted); text-decoration: none; }
.sidebar-footer a:hover { color: var(--accent); }

/* ---------------- Content ---------------- */
.content { padding: 48px 56px; max-width: 960px; }
.page-header { margin-bottom: 32px; padding-bottom: 24px; border-bottom: 1px solid var(--border); }
.page-header h1 {
    font-size: 32px;
    font-weight: 600;
    letter-spacing: -0.02em;
    margin-bottom: 8px;
    word-break: break-word;
}
.page-header .subtitle { color: var(--text-muted); font-size: 16px; }
.page-header .count { color: var(--text-muted); font-size: 14px; margin-top: 4px; }

.breadcrumb { font-size: 13px; color: var(--text-muted); }
.breadcrumb a { color: var(--accent); text-decoration: none; }
.breadcrumb a:hover { text-decoration: underline; }
.breadcrumb span { color: var(--text); }

.external-links {
    margin-top: 32px;
    padding-top: 24px;
    border-top: 1px solid var(--border);
}
.external-links h2 { font-size: 18px; margin-bottom: 16px; }
.link-list { list-style: none; display: flex; flex-wrap: wrap; gap: 12px; }
.link-btn {
    display: inline-flex;
    align-items: center;
    min-height: 44px;
    padding: 12px 20px;
    background: var(--accent);
    color: white;
    text-decoration: none;
    border-radius: var(--radius);
    font-size: 15px;
    font-weight: 500;
    transition: background 0.15s ease;
}
.link-btn:hover, .link-btn:active { background: var(--accent-hover); }

/* ---------------- Acordeones (cards colapsables) ---------------- */
.accordion {
    display: flex;
    flex-direction: column;
    gap: 8px;
    margin-top: 16px;
}

.acc-item {
    background: var(--bg);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    box-shadow: var(--shadow);
    overflow: hidden;
    transition: border-color 0.15s ease, box-shadow 0.15s ease;
}
.acc-item:hover { border-color: var(--accent); }
.acc-item.open { box-shadow: var(--shadow-hover); }

.acc-header {
    display: flex;
    align-items: center;
    gap: 12px;
    width: 100%;
    padding: 16px 18px;
    background: transparent;
    border: 0;
    cursor: pointer;
    text-align: left;
    font-family: inherit;
    color: var(--text);
    min-height: 56px;
}
.acc-header:hover { background: var(--bg-alt); }

.acc-title {
    font-size: 16px;
    font-weight: 600;
    color: var(--text);
    flex: 1;
}
.acc-meta {
    font-size: 13px;
    color: var(--text-muted);
    font-weight: 400;
    margin-right: 4px;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
    max-width: 50%;
}
.acc-chevron {
    font-size: 22px;
    line-height: 1;
    color: var(--text-muted);
    transition: transform 0.2s ease;
    flex-shrink: 0;
}
.acc-item.open .acc-chevron { transform: rotate(90deg); color: var(--accent); }

.acc-body {
    max-height: 0;
    overflow: hidden;
    transition: max-height 0.3s ease;
    padding: 0 18px;
    border-top: 1px solid transparent;
}
.acc-item.open .acc-body {
    max-height: 4000px;
    padding: 14px 18px 18px;
    border-top-color: var(--border);
}

.acc-row {
    display: flex;
    flex-direction: column;
    gap: 2px;
    padding: 8px 0;
    border-bottom: 1px solid var(--border);
}
.acc-row:last-of-type { border-bottom: 0; }
.acc-row-label {
    font-size: 11px;
    font-weight: 600;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.04em;
}
.acc-row-value { font-size: 14px; color: var(--text); }

.acc-list {
    list-style: none;
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(120px, 1fr));
    gap: 6px;
    margin-bottom: 14px;
}
.acc-list li a {
    display: block;
    padding: 8px 10px;
    background: var(--bg-alt);
    border: 1px solid var(--border);
    border-radius: 6px;
    color: var(--text);
    text-decoration: none;
    font-size: 13px;
    font-weight: 500;
    text-align: center;
    transition: background 0.12s ease, border-color 0.12s ease, color 0.12s ease;
}
.acc-list li a:hover, .acc-list li a:active {
    background: rgba(9,105,218,0.08);
    border-color: var(--accent);
    color: var(--accent);
}

.acc-cta {
    display: inline-block;
    margin-top: 8px;
    color: var(--accent);
    text-decoration: none;
    font-size: 14px;
    font-weight: 500;
}
.acc-cta:hover { text-decoration: underline; }

.acc-body .empty {
    color: var(--text-muted);
    font-style: italic;
    font-size: 13px;
    padding: 4px 0 8px;
}
.acc-body p {
    font-size: 15px;
    color: var(--text);
    margin-bottom: 6px;
    line-height: 1.5;
}
.acc-body p:last-child { margin-bottom: 0; }

/* ---------------- Mobile (<= 900px) ---------------- */
@media (max-width: 900px) {
    .layout { grid-template-columns: 1fr; min-height: calc(100vh - var(--banner-h-mobile)); }

    .banner {
        height: var(--banner-h-mobile);
        padding: 0 10px;
        gap: 8px;
    }
    .banner-logo-svg { height: 36px; }
    .banner-title { font-size: 15px; font-weight: 700; }
    .banner-title-full { display: none; }
    .banner-title-short { display: inline; }

    .menu-toggle { display: flex; align-items: center; justify-content: center; width: 40px; height: 40px; font-size: 22px; }
    .search-toggle { width: 40px; height: 40px; }
    .search-modal { padding: 12px; align-items: flex-start; }
    .search-modal-box { max-height: calc(100dvh - 24px); border-radius: 10px; }
    .search-modal-head { padding: 10px 12px; }
    .search-input { font-size: 16px; }  /* evita zoom de iOS */

    .gallery-btn { width: 100%; justify-content: center; padding: 12px; font-size: 15px; }
    .lightbox-prev,
    .lightbox-next { width: 44px; height: 44px; font-size: 32px; }
    .lightbox-prev { left: 8px; }
    .lightbox-next { right: 8px; }
    .lightbox-close { top: 12px; right: 12px; width: 40px; height: 40px; font-size: 20px; }
    .lightbox-img { max-height: calc(100dvh - 130px); }
    .lightbox-counter { bottom: 14px; font-size: 12px; padding: 5px 12px; }

    .sidebar {
        position: fixed;
        top: 0; left: 0;
        width: min(86vw, 320px);
        height: 100dvh;
        z-index: 11;
        transform: translateX(-100%);
        transition: transform 0.25s ease;
        padding: 16px 0 12px;
        box-shadow: 4px 0 16px rgba(0,0,0,0.08);
    }
    .sidebar.open { transform: translateX(0); }

    .brand {
        display: flex;
        align-items: center;
        justify-content: space-between;
        padding: 4px 16px 16px;
    }
    .brand h1 { font-size: 17px; }
    .sidebar-close { display: flex; align-items: center; justify-content: center; }

    .nav-items li a { padding: 12px 14px; font-size: 15px; }

    .content {
        padding: 20px 16px 40px;
        max-width: 100%;
    }
    .page-header { margin-bottom: 20px; padding-bottom: 16px; }
    .page-header h1 { font-size: 26px; }
    .page-header .subtitle { font-size: 15px; }

    .breadcrumb { font-size: 14px; margin-top: 4px; }

    .accordion { gap: 6px; margin-top: 12px; }
    .acc-header { padding: 14px 14px; min-height: 56px; }
    .acc-title { font-size: 15px; }
    .acc-meta { font-size: 12px; max-width: 45%; }
    .acc-item.open .acc-body { padding: 12px 14px 14px; }
    .acc-list { grid-template-columns: repeat(auto-fill, minmax(100px, 1fr)); gap: 6px; }
    .acc-list li a { padding: 10px 8px; font-size: 13px; min-height: 38px; display: flex; align-items: center; justify-content: center; }

    .external-links { margin-top: 24px; padding-top: 18px; }
    .link-list { flex-direction: column; gap: 10px; }
    .link-btn { width: 100%; justify-content: center; font-size: 16px; }
}

@media (max-width: 480px) {
    .banner-title { display: none; }
    .banner { justify-content: space-between; }
    .banner-logo-svg { height: 32px; }
}

@media (max-width: 380px) {
    .page-header h1 { font-size: 22px; }
    .acc-meta { display: none; }
}
"""

SCRIPT_JS = r"""document.addEventListener('DOMContentLoaded', function () {
    var sidebar = document.querySelector('.sidebar');
    var backdrop = document.querySelector('.sidebar-backdrop');
    var toggle = document.querySelector('.menu-toggle');
    var closeBtn = document.querySelector('.sidebar-close');

    function openSidebar() {
        if (!sidebar) return;
        sidebar.classList.add('open');
        if (backdrop) backdrop.classList.add('show');
        if (toggle) toggle.setAttribute('aria-expanded', 'true');
        document.body.style.overflow = 'hidden';
    }
    function closeSidebar() {
        if (!sidebar) return;
        sidebar.classList.remove('open');
        if (backdrop) backdrop.classList.remove('show');
        if (toggle) toggle.setAttribute('aria-expanded', 'false');
        document.body.style.overflow = '';
    }

    if (toggle) toggle.addEventListener('click', openSidebar);
    if (closeBtn) closeBtn.addEventListener('click', closeSidebar);
    if (backdrop) backdrop.addEventListener('click', closeSidebar);

    // Cerrar al navegar (en móvil)
    document.querySelectorAll('.sidebar a').forEach(function (a) {
        a.addEventListener('click', function () {
            if (window.matchMedia('(max-width: 900px)').matches) closeSidebar();
        });
    });

    // Cerrar con tecla Escape
    document.addEventListener('keydown', function (e) {
        if (e.key === 'Escape') closeSidebar();
    });

    // Toggle de categorías
    document.querySelectorAll('.nav-cat').forEach(function (btn) {
        btn.addEventListener('click', function () {
            var group = btn.closest('.nav-group');
            var open = group.classList.toggle('open');
            btn.setAttribute('aria-expanded', open ? 'true' : 'false');
        });
    });

    // Abrir automáticamente la categoría con link activo (o todas si no hay activo)
    var active = document.querySelector('.nav-items a.active');
    if (active) {
        var group = active.closest('.nav-group');
        if (group) {
            group.classList.add('open');
            var b = group.querySelector('.nav-cat');
            if (b) b.setAttribute('aria-expanded', 'true');
        }
    } else {
        document.querySelectorAll('.nav-group').forEach(function (g) {
            g.classList.add('open');
            var b = g.querySelector('.nav-cat');
            if (b) b.setAttribute('aria-expanded', 'true');
        });
    }

    // Scroll al ítem activo dentro del sidebar
    if (active) {
        setTimeout(function () {
            active.scrollIntoView({ block: 'center', behavior: 'instant' });
        }, 30);
    }

    // Acordeones de contenido (cards colapsables)
    document.querySelectorAll('.acc-header').forEach(function (btn) {
        btn.addEventListener('click', function () {
            var item = btn.closest('.acc-item');
            if (!item) return;
            var open = item.classList.toggle('open');
            btn.setAttribute('aria-expanded', open ? 'true' : 'false');
        });
    });

    // ─────────────────────────────────────────────────────────────
    // Búsqueda
    // ─────────────────────────────────────────────────────────────
    var searchToggle = document.querySelector('.search-toggle');
    var searchModal = document.querySelector('.search-modal');
    var searchInput = document.querySelector('.search-input');
    var searchResults = document.querySelector('.search-results');
    var searchClose = document.querySelector('.search-modal-close');
    var searchOverlay = document.querySelector('.search-modal-overlay');
    var rel = document.body.getAttribute('data-rel') || '';

    var searchIndex = null;       // cargado lazy
    var searchLoading = false;
    var activeResultIdx = -1;
    var lastResults = [];

    function escapeHtml(s) {
        return String(s)
            .replace(/&/g, '&amp;')
            .replace(/</g, '&lt;')
            .replace(/>/g, '&gt;')
            .replace(/"/g, '&quot;');
    }

    function highlight(text, query) {
        if (!query) return escapeHtml(text);
        var lower = text.toLowerCase();
        var q = query.toLowerCase();
        var out = '', i = 0, idx;
        while ((idx = lower.indexOf(q, i)) !== -1) {
            out += escapeHtml(text.slice(i, idx));
            out += '<mark>' + escapeHtml(text.slice(idx, idx + q.length)) + '</mark>';
            i = idx + q.length;
        }
        out += escapeHtml(text.slice(i));
        return out;
    }

    function loadIndex() {
        if (searchIndex || searchLoading) return Promise.resolve(searchIndex || []);
        searchLoading = true;
        return fetch(rel + 'search-index.json')
            .then(function (r) { return r.ok ? r.json() : []; })
            .then(function (data) { searchIndex = data; return data; })
            .catch(function () { searchIndex = []; return []; })
            .finally(function () { searchLoading = false; });
    }

    function openSearch() {
        if (!searchModal) return;
        searchModal.hidden = false;
        document.body.classList.add('search-open');
        searchToggle && searchToggle.setAttribute('aria-expanded', 'true');
        loadIndex().then(function () {
            // foco al input después del paint para que en móvil aparezca el teclado
            setTimeout(function () { searchInput && searchInput.focus(); }, 20);
            renderResults(searchInput.value);
        });
    }

    function closeSearch() {
        if (!searchModal) return;
        searchModal.hidden = true;
        document.body.classList.remove('search-open');
        searchToggle && searchToggle.setAttribute('aria-expanded', 'false');
        activeResultIdx = -1;
    }

    function renderResults(query) {
        if (!searchResults) return;
        query = (query || '').trim();
        if (!searchIndex) {
            searchResults.innerHTML = '<div class="search-empty">Cargando índice…</div>';
            return;
        }
        if (!query) {
            // Sin query: mostrar todos (hasta 50)
            lastResults = searchIndex.slice(0, 50);
        } else {
            var q = query.toLowerCase();
            // Filtrar por substring en el haystack
            var matches = [];
            for (var i = 0; i < searchIndex.length; i++) {
                if (searchIndex[i].haystack.indexOf(q) !== -1) {
                    matches.push(searchIndex[i]);
                }
            }
            // Ordenar: TAGs que empiezan con la query primero, luego por TAG alfabético
            matches.sort(function (a, b) {
                var aStarts = a.tag.toLowerCase().indexOf(q) === 0 ? 0 : 1;
                var bStarts = b.tag.toLowerCase().indexOf(q) === 0 ? 0 : 1;
                if (aStarts !== bStarts) return aStarts - bStarts;
                return a.tag.localeCompare(b.tag);
            });
            lastResults = matches.slice(0, 50);
        }

        if (lastResults.length === 0) {
            searchResults.innerHTML = '<div class="search-empty">Sin resultados para “' + escapeHtml(query) + '”.</div>';
            activeResultIdx = -1;
            return;
        }

        var html = '';
        for (var j = 0; j < lastResults.length; j++) {
            var r = lastResults[j];
            html += '<a class="search-result" href="' + rel + r.url + '" data-idx="' + j + '">';
            html += '<div class="search-result-head">';
            html += '<span class="search-result-tag">' + highlight(r.tag, query) + '</span>';
            html += '<span class="search-result-cat">' + escapeHtml(r.category) + '</span>';
            html += '</div>';
            if (r.preview) {
                html += '<div class="search-result-preview">' + highlight(r.preview, query) + '</div>';
            }
            html += '</a>';
        }
        searchResults.innerHTML = html;
        activeResultIdx = lastResults.length ? 0 : -1;
        updateActiveResult();
    }

    function updateActiveResult() {
        var nodes = searchResults.querySelectorAll('.search-result');
        nodes.forEach(function (n, i) {
            n.classList.toggle('active', i === activeResultIdx);
        });
        if (activeResultIdx >= 0 && nodes[activeResultIdx]) {
            nodes[activeResultIdx].scrollIntoView({ block: 'nearest' });
        }
    }

    // Debounce del input
    var searchTimer = null;
    if (searchInput) {
        searchInput.addEventListener('input', function () {
            clearTimeout(searchTimer);
            searchTimer = setTimeout(function () {
                renderResults(searchInput.value);
            }, 80);
        });
        // Flechas + Enter para navegar resultados
        searchInput.addEventListener('keydown', function (e) {
            if (e.key === 'ArrowDown') {
                e.preventDefault();
                if (lastResults.length === 0) return;
                activeResultIdx = (activeResultIdx + 1) % lastResults.length;
                updateActiveResult();
            } else if (e.key === 'ArrowUp') {
                e.preventDefault();
                if (lastResults.length === 0) return;
                activeResultIdx = (activeResultIdx - 1 + lastResults.length) % lastResults.length;
                updateActiveResult();
            } else if (e.key === 'Enter') {
                e.preventDefault();
                if (activeResultIdx >= 0 && lastResults[activeResultIdx]) {
                    window.location.href = rel + lastResults[activeResultIdx].url;
                }
            }
        });
    }

    if (searchToggle) searchToggle.addEventListener('click', openSearch);
    if (searchClose) searchClose.addEventListener('click', closeSearch);
    if (searchOverlay) searchOverlay.addEventListener('click', closeSearch);

    // Esc cierra (sobreescribe el handler genérico de Esc para sidebar si la búsqueda está abierta)
    document.addEventListener('keydown', function (e) {
        if (e.key === 'Escape' && searchModal && !searchModal.hidden) {
            e.stopPropagation();
            closeSearch();
        }
    }, true);  // capture phase para correr antes del Esc del sidebar

    // ─────────────────────────────────────────────────────────────
    // Lightbox / galería de imágenes
    // ─────────────────────────────────────────────────────────────
    var lightbox = document.getElementById('lightbox');
    var lbImg = lightbox && lightbox.querySelector('.lightbox-img');
    var lbCaption = lightbox && lightbox.querySelector('.lightbox-caption');
    var lbCounter = lightbox && lightbox.querySelector('.lightbox-counter');
    var lbPrev = lightbox && lightbox.querySelector('.lightbox-prev');
    var lbNext = lightbox && lightbox.querySelector('.lightbox-next');
    var lbClose = lightbox && lightbox.querySelector('.lightbox-close');
    var lbOverlay = lightbox && lightbox.querySelector('.lightbox-overlay');
    var lbImages = [];
    var lbCurrent = 0;
    var lbTag = '';
    var lbTouchStartX = null;

    function lbRender() {
        if (!lbImages.length) return;
        var url = lbImages[lbCurrent];
        lbImg.src = url;
        lbImg.alt = lbTag + ' — imagen ' + (lbCurrent + 1);
        lbCaption.textContent = lbTag + ' — ' + (lbCurrent + 1) + ' / ' + lbImages.length;
        lbCounter.textContent = (lbCurrent + 1) + ' / ' + lbImages.length;
        lightbox.classList.toggle('single', lbImages.length <= 1);
    }
    function lbOpen(images, tag) {
        if (!lightbox || !images || !images.length) return;
        lbImages = images.slice();
        lbCurrent = 0;
        lbTag = tag || '';
        lbRender();
        lightbox.hidden = false;
        document.body.classList.add('lightbox-open');
    }
    function lbCloseFn() {
        if (!lightbox) return;
        lightbox.hidden = true;
        document.body.classList.remove('lightbox-open');
        lbImg.src = '';  // libera memoria
    }
    function lbStep(delta) {
        if (lbImages.length <= 1) return;
        lbCurrent = (lbCurrent + delta + lbImages.length) % lbImages.length;
        lbRender();
    }

    // Conectar los botones "Ver imágenes" de cada página
    document.querySelectorAll('.gallery-btn').forEach(function (btn) {
        btn.addEventListener('click', function () {
            var raw = btn.getAttribute('data-images') || '[]';
            try {
                var imgs = JSON.parse(raw);
                lbOpen(imgs, btn.getAttribute('data-tag') || '');
            } catch (e) { /* ignore */ }
        });
    });

    if (lbPrev) lbPrev.addEventListener('click', function () { lbStep(-1); });
    if (lbNext) lbNext.addEventListener('click', function () { lbStep(1); });
    if (lbClose) lbClose.addEventListener('click', lbCloseFn);
    if (lbOverlay) lbOverlay.addEventListener('click', lbCloseFn);

    // Teclado: ← → para navegar, Esc para cerrar
    document.addEventListener('keydown', function (e) {
        if (!lightbox || lightbox.hidden) return;
        if (e.key === 'ArrowLeft') { e.preventDefault(); lbStep(-1); }
        else if (e.key === 'ArrowRight') { e.preventDefault(); lbStep(1); }
        else if (e.key === 'Escape') { e.stopPropagation(); lbCloseFn(); }
    }, true);

    // Swipe touch en móvil
    if (lightbox) {
        lightbox.addEventListener('touchstart', function (e) {
            lbTouchStartX = e.touches[0].clientX;
        }, { passive: true });
        lightbox.addEventListener('touchend', function (e) {
            if (lbTouchStartX === null) return;
            var dx = e.changedTouches[0].clientX - lbTouchStartX;
            if (Math.abs(dx) > 50) lbStep(dx < 0 ? 1 : -1);
            lbTouchStartX = null;
        }, { passive: true });
    }
});
"""


# ─────────────────────────────────────────────────────────────────────────
# Generación del sitio
# ─────────────────────────────────────────────────────────────────────────
def generate_site(items: list[Item], output_dir: Path, config: Optional[dict] = None):
    print(f"🏗️  Generando sitio en: {output_dir}")
    if config is None:
        config = load_site_config(SITE_CONFIG_PATH)
    grouped = group_by_category(items)

    # Limpiar docs/ preservando docs/assets/ (imágenes y otros recursos estáticos)
    if output_dir.exists():
        for entry in output_dir.iterdir():
            if entry.name == "assets":
                continue
            if entry.is_dir():
                shutil.rmtree(entry)
            else:
                entry.unlink()
    else:
        output_dir.mkdir(parents=True)

    # Assets estáticos (CSS = defaults + override de site_config.json)
    final_css = STYLES_CSS + build_theme_override_css(config["theme"])
    (output_dir / "styles.css").write_text(final_css, encoding="utf-8")
    (output_dir / "script.js").write_text(SCRIPT_JS, encoding="utf-8")
    (output_dir / ".nojekyll").write_text("", encoding="utf-8")

    # Icono genérico (fallback del thumbnail del botón "Ver imagen")
    icons_dir = output_dir / "assets" / "icons"
    icons_dir.mkdir(parents=True, exist_ok=True)
    (icons_dir / "generic-image.svg").write_text(GENERIC_IMAGE_SVG, encoding="utf-8")
    print("   ✅ styles.css + script.js + .nojekyll + assets/icons/generic-image.svg")

    # Home
    sidebar_home = build_nav(grouped, rel_prefix="")
    home_html = page_skeleton(
        page_title="Inicio",
        topbar_title="Inicio",
        sidebar_html=sidebar_home,
        content_html=render_home(grouped),
        rel_prefix="",
    )
    (output_dir / "index.html").write_text(home_html, encoding="utf-8")
    print("   ✅ index.html")

    # Páginas por categoría y por item
    total_items = 0
    for cat, cat_items in grouped.items():
        if not cat_items:
            continue
        cat_dir = output_dir / cat.lower()
        cat_dir.mkdir(exist_ok=True)

        # Index de categoría
        sidebar_cat = build_nav(grouped, rel_prefix="../")
        cat_index_html = page_skeleton(
            page_title=cat,
            topbar_title=cat,
            sidebar_html=sidebar_cat,
            content_html=render_category_index(cat, cat_items),
            rel_prefix="../",
        )
        (cat_dir / "index.html").write_text(cat_index_html, encoding="utf-8")

        # Páginas individuales
        for item in cat_items:
            sidebar_item = build_nav(grouped, rel_prefix="../", active_tag=item.tag)
            item_html = page_skeleton(
                page_title=item.title,
                topbar_title=item.tag,
                sidebar_html=sidebar_item,
                content_html=render_item_page(item),
                rel_prefix="../",
            )
            (cat_dir / item.filename).write_text(item_html, encoding="utf-8")
            total_items += 1

        print(f"   ✅ {cat.lower()}/ ({len(cat_items)} páginas)")

    # README + metadata
    readme = (
        f"# {SITE_TITLE}\n\n"
        f"Sitio generado automáticamente desde `plantilla_sitio.xlsx`.\n\n"
        f"- **Fecha:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"- **TAGs publicados:** {total_items}\n"
        f"- **Categorías:** {', '.join(c for c in CATEGORY_ORDER if grouped.get(c))}\n\n"
        f"URL pública: {SITE_URL}\n"
    )
    (output_dir / "README.md").write_text(readme, encoding="utf-8")

    metadata = {
        "source": "plantilla_sitio.xlsx",
        "generated_at": datetime.now().isoformat(),
        "site_title": SITE_TITLE,
        "site_url": SITE_URL,
        "total_tags": total_items,
        "categories": {c: len(grouped.get(c, [])) for c in CATEGORY_ORDER},
        "tags": [
            {
                "tag": i.tag,
                "category": i.category,
                "title": i.title,
                "properties": i.properties,
                "ficha_tecnica_url": i.ficha_tecnica_url,
                "curva_url": i.curva_url,
                "resources": i.resources,
            }
            for i in items
        ],
    }
    (output_dir / "migration_metadata.json").write_text(
        json.dumps(metadata, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    # Índice de búsqueda (cargado por search.js en el cliente)
    search_index = []
    for it in items:
        # Texto plano de todo el contenido para hacer búsquedas substring
        haystack_parts = [it.tag, it.category, it.title]
        for k, v in it.properties.items():
            haystack_parts.append(k)
            haystack_parts.append(v)
        haystack = " ".join(haystack_parts).lower()
        search_index.append({
            "tag": it.tag,
            "category": it.category,
            "title": it.title,
            "url": f"{it.category.lower()}/{it.filename}",
            "haystack": haystack,
            # Vista previa: primeras 2 propiedades con valor
            "preview": " · ".join(
                f"{k}: {v[:60]}" for k, v in list(it.properties.items())[:2]
            ),
        })
    (output_dir / "search-index.json").write_text(
        json.dumps(search_index, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print("   ✅ README.md + migration_metadata.json + search-index.json")
    print()
    print(f"✅ {total_items} TAGs publicados en {output_dir}/")
    print(f"   Por categoría: " + ", ".join(
        f"{c}={len(grouped.get(c, []))}" for c in CATEGORY_ORDER if grouped.get(c)
    ))


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────
def main():
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT

    if not input_path.exists():
        print(f"❌ No existe: {input_path}", file=sys.stderr)
        sys.exit(1)

    # Cargar config y aplicar overrides de título/URL/color a los globals
    config = load_site_config(SITE_CONFIG_PATH)
    global SITE_TITLE, BANNER_TITLE_FULL, BANNER_TITLE_SHORT, SITE_URL, _MOBILE_THEME_COLOR
    SITE_TITLE = config["site_title"]
    BANNER_TITLE_FULL = config["banner_title_full"]
    BANNER_TITLE_SHORT = config["banner_title_short"]
    SITE_URL = config["site_url"]
    _MOBILE_THEME_COLOR = config["theme"].get("accent", _MOBILE_THEME_COLOR)

    items, _ = load_items(input_path)
    if not items:
        print("❌ No hay TAGs en la plantilla", file=sys.stderr)
        sys.exit(1)

    generate_site(items, output_dir, config=config)


if __name__ == "__main__":
    main()
