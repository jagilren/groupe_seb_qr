#!/usr/bin/env python3
"""
convert_fichas_to_template.py
=============================
Convierte un Excel de fichas técnicas (una hoja por TAG) en la plantilla
unificada compatible con el migrador del sitio (una sola hoja con todos
los TAGs en filas y propiedades en columnas).

Input:  NP00033 Fichas Técnicas Equipos 20250408 DG.xlsx (o similar)
Output: plantilla_sitio.xlsx (en la raíz del proyecto)

Reglas:
- Cada hoja del Excel origen es uno o varios TAGs.
- TAGs múltiples se expanden:
    "100-P-01 A_B"          -> "100-P-01A", "100-P-01B"
    "400-MX-01_02_03_04"    -> "400-MX-01", "400-MX-02", "400-MX-03", "400-MX-04"
- Categoría inferida del tipo en el TAG:
    TK, R, F, SD            -> TANQUES
    P, MX, M, C, PS         -> EQUIPOS
    AIT, FIT, SV, LT        -> INSTRUMENTOS
- Cada propiedad detectada en el archivo origen se vuelve una columna.
- Celdas vacías se mantienen vacías (no se genera esa sección en el sitio).
- Columnas 'Ficha Técnica' y 'Curva' se incluyen pero quedan vacías
  (este archivo no aporta links a documentos).
"""

import re
import shutil
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────────────────────────────────
# Configuración
# ─────────────────────────────────────────────────────────────────────────
REPO_ROOT = Path(__file__).parent
INPUT_DEFAULT = REPO_ROOT / "NP00033 Fichas Técnicas Equipos 20250408 DG.xlsx"
OUTPUT = REPO_ROOT / "plantilla_sitio.xlsx"
IMAGES_DIR = REPO_ROOT / "docs" / "assets" / "images"

# Imágenes que aparecen en MÁS de este umbral de hojas se consideran plantilla
# (logo, header, etc.) y se descartan. Solo se conservan las específicas del equipo.
IMAGE_COMMON_THRESHOLD = 5

CATEGORIES = ["EQUIPOS", "INSTRUMENTOS", "TANQUES"]

# Mapping de prefijo de tipo (tras el primer número) → categoría
TYPE_TO_CATEGORY = {
    # TANQUES
    "TK": "TANQUES",
    "R": "TANQUES",
    "F": "TANQUES",
    "SD": "TANQUES",
    # EQUIPOS
    "P": "EQUIPOS",
    "MX": "EQUIPOS",
    "M": "EQUIPOS",
    "C": "EQUIPOS",
    "PS": "EQUIPOS",
    # INSTRUMENTOS
    "AIT": "INSTRUMENTOS",
    "FIT": "INSTRUMENTOS",
    "SV": "INSTRUMENTOS",
    "LT": "INSTRUMENTOS",
}

# Labels a IGNORAR (metadata administrativa, headers de sección)
IGNORE_LABELS = {
    # Headers de bloque
    "GENERAL", "TANQUE", "EQUIPO", "FLUIDO", "NOTAS", "FICHA TÉCNICA",
    # Metadata administrativa
    "PLANTA", "UBICACIÓN", "VERSIÓN", "FECHA", "ELABORÓ", "REVISÓ", "SOPORTE",
    # Constantes redundantes que aparecen pegadas
    "STARND ETAPA 1 Y 2", "ITAGUI - ANTIOQUIA",
    # Etiqueta del TAG (lo manejamos por separado)
    "TAG N°", "TAG N",
}


# ─────────────────────────────────────────────────────────────────────────
# Helpers de TAG
# ─────────────────────────────────────────────────────────────────────────
def expand_tags_from_sheet_name(sheet_name: str) -> list[str]:
    """Expande nombres de hoja a una lista de TAGs individuales."""
    name = sheet_name.strip()

    # Patrón "XXX A_B" o "XXX A/B" → XXXA, XXXB
    m = re.match(r"^(.+?)\s*A[_/]B\s*$", name)
    if m:
        base = m.group(1).strip()
        return [f"{base}A", f"{base}B"]

    # Patrón "XXX-NN_MM_OO_PP" → XXX-NN, XXX-MM, ...
    m = re.match(r"^(.+?-)(\d+(?:_\d+)+)$", name)
    if m:
        prefix, nums = m.group(1), m.group(2)
        return [f"{prefix}{n}" for n in nums.split("_")]

    # Sin patrón → TAG único
    return [name]


def infer_category(tag: str) -> str:
    """Deriva la categoría de un TAG por su tipo (XXX-TYPE-NN)."""
    parts = tag.split("-")
    if len(parts) < 2:
        return "EQUIPOS"  # fallback
    # parts[1] suele ser el tipo (TK, P, MX, M, C, PS, AIT, FIT, SV, LT, R, F, SD)
    type_code = parts[1].rstrip("0123456789 ").strip()
    return TYPE_TO_CATEGORY.get(type_code, "EQUIPOS")


# ─────────────────────────────────────────────────────────────────────────
# Extracción de pares (label, value)
# ─────────────────────────────────────────────────────────────────────────
def normalize_label(label: str) -> str:
    """Limpia espacios duplicados y caracteres raros del label."""
    s = re.sub(r"\s+", " ", label.strip())
    return s


def is_valid_label(label: str) -> bool:
    """Filtra labels que no son propiedades reales (metadata, headers, etc.)."""
    if not label or len(label) < 2 or len(label) > 80:
        return False
    upper = label.upper().strip()
    if upper in IGNORE_LABELS:
        return False
    # Texto que es una fecha
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", label):
        return False
    # Si es una etiqueta tipo "Versión"/"Fecha"/"Elaboró"
    if upper in ("VERSIÓN", "FECHA", "ELABORÓ", "REVISÓ"):
        return False
    return True


def format_value(value) -> str:
    """Convierte un valor de celda a string limpio."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def extract_pairs(ws) -> dict[str, str]:
    """
    Devuelve un dict {label: value} con las propiedades de la hoja.
    Recorre fila por fila buscando pares de celdas adyacentes label:value.
    Si el mismo label aparece dos veces, conserva el primer valor no vacío.
    """
    pairs: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        i = 0
        while i < len(cells) - 1:
            cell = cells[i]
            if cell is None or not isinstance(cell, str):
                i += 1
                continue
            label = normalize_label(cell)
            # Buscar el primer valor no vacío en las celdas adyacentes (hasta 3)
            value = None
            j = i + 1
            while j < len(cells) and j <= i + 3:
                v = cells[j]
                if v is not None and format_value(v):
                    value = v
                    break
                j += 1
            if value is not None and is_valid_label(label):
                value_str = format_value(value)
                # El valor no puede ser un label conocido (evita capturar headers como valores)
                if value_str.upper() not in IGNORE_LABELS and label not in pairs:
                    pairs[label] = value_str
            i += 1
    return pairs


# ─────────────────────────────────────────────────────────────────────────
# Extracción de imágenes embebidas en el Excel
# ─────────────────────────────────────────────────────────────────────────
_XLSX_NS = {
    'r':    'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
}


def _resolve_target(base: str, target: str) -> str:
    """Resuelve una ruta relativa estilo '../media/image1.png' contra una base."""
    target = target.lstrip('/')
    if target.startswith('../'):
        # Subir niveles desde 'base' (que es la carpeta del archivo origen)
        base_parts = base.split('/')[:-1]  # carpeta
        while target.startswith('../'):
            target = target[3:]
            if base_parts:
                base_parts.pop()
        return '/'.join(base_parts + [target])
    return '/'.join(base.split('/')[:-1] + [target])


def _sheet_to_images_map(zf: zipfile.ZipFile) -> dict[str, list[str]]:
    """Devuelve {sheet_name: [paths de imagen en el zip]} para cada hoja."""
    wb_xml = ET.fromstring(zf.read('xl/workbook.xml'))
    sheets = [{'name': s.get('name'),
               'rid': s.get('{%s}id' % _XLSX_NS['r'])}
              for s in wb_xml.find('main:sheets', _XLSX_NS)]
    wb_rels = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
    rid_to_target = {r.get('Id'): r.get('Target') for r in wb_rels}

    result: dict[str, list[str]] = {}
    for s in sheets:
        sheet_path = 'xl/' + rid_to_target[s['rid']]
        rels_path = sheet_path.replace('worksheets/', 'worksheets/_rels/') + '.rels'
        if rels_path not in zf.namelist():
            result[s['name']] = []
            continue
        sheet_rels = ET.fromstring(zf.read(rels_path))
        drawing_target = next(
            (r.get('Target') for r in sheet_rels
             if 'drawing' in r.get('Type', '').lower()),
            None,
        )
        if not drawing_target:
            result[s['name']] = []
            continue
        drawing_path = _resolve_target(sheet_path, drawing_target)
        drawing_rels_path = (
            drawing_path.replace('drawings/', 'drawings/_rels/') + '.rels'
        )
        if drawing_rels_path not in zf.namelist():
            result[s['name']] = []
            continue
        drawing_rels = ET.fromstring(zf.read(drawing_rels_path))
        images = []
        for r in drawing_rels:
            if 'image' in r.get('Type', '').lower():
                images.append(_resolve_target(drawing_path, r.get('Target')))
        result[s['name']] = images
    return result


def extract_images(input_path: Path, images_dir: Path) -> dict[str, list[str]]:
    """
    Extrae las imágenes específicas de cada hoja del Excel origen a
    images_dir/<TAG>/N.<ext> y devuelve un dict {TAG: [filenames]}.

    Las imágenes comunes (plantilla, logos, headers) se descartan
    automáticamente: aparecen en más de IMAGE_COMMON_THRESHOLD hojas.
    """
    print(f"🖼️  Extrayendo imágenes a: {images_dir}")
    images_dir.mkdir(parents=True, exist_ok=True)

    extracted: dict[str, list[str]] = {}
    with zipfile.ZipFile(input_path) as zf:
        sheet_to_images = _sheet_to_images_map(zf)

        # Contar en cuántas hojas aparece cada imagen → descartar plantillas
        usage = defaultdict(int)
        for imgs in sheet_to_images.values():
            for img in imgs:
                usage[img] += 1

        for sheet_name, image_paths in sheet_to_images.items():
            specific = [p for p in image_paths
                       if usage[p] <= IMAGE_COMMON_THRESHOLD]
            if not specific:
                continue
            # Expandir el nombre de hoja a TAGs (igual que en process_workbook)
            tags = expand_tags_from_sheet_name(sheet_name)
            for tag in tags:
                # Slug del TAG (mismo que en excel_migrator.Item.filename)
                slug = tag.replace(' ', '').replace('/', '_')
                tag_dir = images_dir / slug
                tag_dir.mkdir(parents=True, exist_ok=True)
                filenames: list[str] = []
                for i, img_path in enumerate(specific, start=1):
                    if img_path not in zf.namelist():
                        continue
                    ext = Path(img_path).suffix.lower()  # .png / .jpeg
                    out_name = f"{i}{ext}"
                    out_file = tag_dir / out_name
                    with zf.open(img_path) as src, open(out_file, 'wb') as dst:
                        shutil.copyfileobj(src, dst)
                    filenames.append(out_name)
                if filenames:
                    extracted[slug] = filenames

    # Escribir manifest para que el migrador lo lea
    manifest = {tag: files for tag, files in sorted(extracted.items())}
    manifest_path = images_dir / "manifest.json"
    import json
    manifest_path.write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"   ✅ {len(extracted)} TAGs con imágenes, manifest.json escrito")
    return extracted


# ─────────────────────────────────────────────────────────────────────────
# Procesamiento del Excel origen
# ─────────────────────────────────────────────────────────────────────────
def process_workbook(input_path: Path) -> tuple[list[dict], list[str]]:
    """
    Lee el Excel de fichas y devuelve:
        - rows: lista de dicts {TAG, Categoría, Título, <propiedades>}
        - prop_columns: lista ordenada de columnas de propiedades detectadas
    """
    print(f"📖 Leyendo: {input_path}")
    wb = load_workbook(input_path, data_only=True)

    rows: list[dict] = []
    all_props: list[str] = []  # mantiene el orden de aparición
    seen_props: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tags = expand_tags_from_sheet_name(sheet_name)
        pairs = extract_pairs(ws)
        category = infer_category(tags[0])

        # Acumular orden de aparición de las propiedades (preserva orden del Excel origen)
        for k in pairs:
            if k not in seen_props:
                seen_props.add(k)
                all_props.append(k)

        for tag in tags:
            row = {"TAG": tag, "Categoría": category, "Título": ""}
            row.update(pairs)
            rows.append(row)

        print(f"  ✓ {sheet_name:30} → {len(tags)} TAG(s) [{category:13}] · {len(pairs)} propiedades")

    return rows, all_props


# ─────────────────────────────────────────────────────────────────────────
# Generación del Excel destino
# ─────────────────────────────────────────────────────────────────────────
def style_header(ws, n_cols):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="0969DA")
    border = Border(bottom=Side(style="medium", color="0550AE"))
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 36


def autosize_columns(ws, headers, rows):
    for i, h in enumerate(headers, start=1):
        max_len = max(len(h), 12)
        for r in rows:
            v = str(r.get(h, "") or "")
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 3, 14), 50)


def add_category_validation(ws, n_rows):
    formula = f'"{",".join(CATEGORIES)}"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=False,
        showDropDown=False,
        errorTitle="Categoría inválida",
        error=f"Debe ser una de: {', '.join(CATEGORIES)}",
        promptTitle="Categoría",
        prompt=f"Selecciona: {', '.join(CATEGORIES)}",
    )
    ws.add_data_validation(dv)
    last_row = max(n_rows + 100, 500)
    dv.add(f"B2:B{last_row}")


def write_data_sheet(wb: Workbook, rows: list[dict], prop_columns: list[str]):
    ws = wb.active
    ws.title = "Datos"

    headers = ["TAG", "Categoría", "Título"] + prop_columns + ["Ficha Técnica", "Curva"]

    # Headers
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    style_header(ws, len(headers))

    # Filas
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, h in enumerate(headers, start=1):
            v = row.get(h, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=v if v else None)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "C2"
    add_category_validation(ws, len(rows))
    autosize_columns(ws, headers, rows)


def write_instructions_sheet(wb: Workbook):
    ws = wb.create_sheet("Instrucciones")
    title_font = Font(bold=True, size=14, color="0969DA")
    h2_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    code_font = Font(size=10, name="Consolas", color="333333")

    lines = [
        ("Plantilla de datos para el sitio QR Groupe SEB", title_font),
        ("", normal_font),
        ("Cada fila es un TAG. Cada columna a partir de 'Título' es una propiedad", normal_font),
        ("que aparecerá como sección colapsable en la ficha del TAG.", normal_font),
        ("Si dejas una celda vacía, esa sección NO se genera en la ficha.", normal_font),
        ("", normal_font),
        ("Columnas obligatorias", h2_font),
        ("• TAG          identificador único (ej. 100-P-01A)", normal_font),
        ("• Categoría    EQUIPOS, INSTRUMENTOS o TANQUES (dropdown en columna B)", normal_font),
        ("", normal_font),
        ("Columnas opcionales", h2_font),
        ("• Título           si lo dejas vacío, se usa el TAG como título", normal_font),
        ("• <propiedades>    cada columna detectada del Excel origen", normal_font),
        ("• Ficha Técnica    inserta hyperlink con Ctrl+K → pega URL de Drive", normal_font),
        ("• Curva            ídem", normal_font),
        ("", normal_font),
        ("Cómo regenerar este Excel desde fichas técnicas", h2_font),
        ("Si recibes un Excel con una hoja por TAG (formato ficha técnica), ejecuta:", normal_font),
        ("    python3 convert_fichas_to_template.py <archivo.xlsx>", code_font),
        ("Se sobrescribirá esta plantilla con los datos del nuevo archivo.", normal_font),
        ("", normal_font),
        ("Cómo regenerar el sitio desde este Excel", h2_font),
        ("    ./update_site_from_excel.sh         # regenera docs/ sin pushear", code_font),
        ("    ./update_site_from_excel.sh --push  # regenera y publica a GitHub Pages", code_font),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = font
        c.alignment = Alignment(vertical="top")
    ws.column_dimensions["A"].width = 95
    ws.sheet_view.showGridLines = False


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────
def main():
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else INPUT_DEFAULT
    if not input_path.exists():
        print(f"❌ No existe: {input_path}", file=sys.stderr)
        sys.exit(1)

    rows, prop_columns = process_workbook(input_path)

    # Extraer imágenes embebidas del Excel origen a docs/assets/images/
    print()
    extract_images(input_path, IMAGES_DIR)
    print()

    wb = Workbook()
    write_data_sheet(wb, rows, prop_columns)
    write_instructions_sheet(wb)
    wb.save(OUTPUT)

    print()
    print(f"✅ Plantilla generada: {OUTPUT}")
    print(f"   • {len(rows)} TAGs en total")
    print(f"   • {len(prop_columns)} columnas de propiedades + Ficha Técnica + Curva")
    print()
    print("Categorías:")
    from collections import Counter
    counts = Counter(r["Categoría"] for r in rows)
    for cat in CATEGORIES:
        print(f"  {cat:13} {counts.get(cat, 0)}")


if __name__ == "__main__":
    main()
